#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon May 10 16:25:21 2021

@author: virati
Filter through patient list for Lung Mets Specifically

"""

import pandas as pd
import openpyxl as opyx
from openpyxl import Workbook
from collections import defaultdict
import collections
import datetime
from operator import getitem
import string

''' Auto nested dictionary '''
def nestdict():
    return defaultdict(nestdict)


#%%
wb = opyx.load_workbook('../data/brain_mets.xlsx')
#pd.read_excel('../data/brain_mets.xlsx')
ws = wb.active
#%%

charts = []
dcharts = nestdict()

for row in ws.iter_rows(min_row = 4, max_col = 7, max_row=4432):#ws['B4':'B4432']:
    pt = row[1].value
    #surgeon
    
    if row[0].value == None:
        surgeon = active_surgeon
    else:
        active_surgeon = row[0].value
    
    if pt != None:
        active_pt = pt
        dcharts[pt]['EMPI'] = row[2].value
    else:
        pt = active_pt
    dcharts[pt]['Last Name'] = pt.split()[0][:-1]
    dcharts[pt]['First Name'] = pt.split()[1]

    if row[4].value != None:
        dcharts[pt]['Dx'] = row[4].value
    
    dcharts[pt]['Pathology'] = '?'
    
    if dcharts[pt]['First Date']:
        if row[6].value != None and row[6].value < dcharts[pt]['First Date']:
            dcharts[pt]['First Date'] = row[6].value
    else:
        dcharts[pt]['First Date'] = row[6].value
    dcharts[pt]['Surgeon'] = active_surgeon
    dcharts[pt]['Surgery Date'] = []
    dcharts[pt]['Alk'] = []
    dcharts[pt]['EGFR'] = []

    if dcharts[pt]['Dates']:
        dcharts[pt]['Dates'].append(row[6].value)
    else:
        dcharts[pt]['Dates'] = [row[6].value]
    
        
#%%
print(str(len(dcharts)) + ' total patients.')


#%%
#Sort through the dates now
date_dcharts = {key: val for key,val in dcharts.items() if val['First Date'] < datetime.datetime(2019,12,31,0,0) and val['First Date'] > datetime.datetime(2012,1,1,0,0)}

#%%
import xlsxwriter
#date sorted dictionary
sorted_dcharts = collections.OrderedDict(sorted(date_dcharts.items(),key=lambda x: getitem(x[1],'First Date')))


out_frame = pd.DataFrame(sorted_dcharts).T

writer = pd.ExcelWriter('../data/pt_list.xlsx', engine='xlsxwriter')
out_frame.to_excel(writer,sheet_name='Sheet1')
writer.save()